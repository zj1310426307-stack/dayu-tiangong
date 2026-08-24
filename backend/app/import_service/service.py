"""解析导入文件、校验行数据并执行原子批量写入。"""

import csv
import io
import json
from datetime import UTC, datetime
from pathlib import Path
from typing import Any, Literal
from uuid import uuid4

from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.cross_section import service as cross_section_service
from app.dataset.lifecycle import assert_dataset_version_mutable
from app.cross_section.schemas import CrossSectionCreate
from app.files import atomic_write_bytes, storage_directory
from app.gis.models import River
from app.import_service.schemas import ImportIssue, ImportResponse
from app.river import service as river_service
from app.river.schemas import RiverCreate
from app.structure import service as structure_service
from app.structure.schemas import GateCreate, PumpCreate


ResourceName = Literal["rivers", "cross_sections", "gates", "pumps"]
STORAGE_ROOT = storage_directory("imports")


def store_upload(filename: str, content: bytes) -> str:
    """把原始上传文件存档到项目专属目录并返回安全文件名。"""

    safe_suffix = Path(filename).suffix.lower()[:10]
    timestamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")
    stored_name = f"{timestamp}_{uuid4().hex[:10]}{safe_suffix}"
    atomic_write_bytes(STORAGE_ROOT, stored_name, content)
    return stored_name


def parse_excel(content: bytes) -> list[dict[str, Any]]:
    """读取活动工作表，以首行作为字段名返回记录。"""

    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    records = [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        for row in rows
        if any(value is not None and value != "" for value in row)
    ]
    workbook.close()
    return records


def parse_csv(content: bytes) -> list[dict[str, Any]]:
    """按 UTF-8-SIG 解析 CSV 表头和记录。"""

    text = content.decode("utf-8-sig")
    return [dict(row) for row in csv.DictReader(io.StringIO(text))]


def parse_geojson(content: bytes) -> list[dict[str, Any]]:
    """解析 GeoJSON FeatureCollection 并合并 properties 与 geometry。"""

    payload = json.loads(content.decode("utf-8-sig"))
    if payload.get("type") != "FeatureCollection" or not isinstance(payload.get("features"), list):
        raise ValueError("GeoJSON 根对象必须是 FeatureCollection")
    records: list[dict[str, Any]] = []
    for feature in payload["features"]:
        if not isinstance(feature, dict) or feature.get("type") != "Feature":
            raise ValueError("features 中存在非 Feature 对象")
        record = dict(feature.get("properties") or {})
        record["geometry"] = feature.get("geometry")
        records.append(record)
    return records


def _json_value(value: Any, field: str) -> Any:
    """兼容模板中的 JSON 文本与 GeoJSON 原生对象。"""

    if isinstance(value, (dict, list)):
        return value
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"{field} 不能为空")
    return json.loads(value)


def _float(value: Any, field: str) -> float:
    """把表格单元格转换为浮点数并提供字段化错误。"""

    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field} 必须是数值") from exc


def _river_id(session: Session, dataset_version_id: int, river_code: Any) -> int:
    """按数据版本和河道编码解析外键。"""

    river = session.scalar(
        select(River).where(
            River.dataset_version_id == dataset_version_id,
            River.code == str(river_code).strip(),
        )
    )
    if river is None:
        raise ValueError(f"river_code={river_code!s} 在该版本不存在")
    return river.id


def _point(record: dict[str, Any]) -> dict[str, Any]:
    """从 GeoJSON geometry 或经纬度列构造 Point。"""

    if record.get("geometry"):
        return record["geometry"]
    return {
        "type": "Point",
        "coordinates": [_float(record.get("longitude"), "longitude"), _float(record.get("latitude"), "latitude")],
    }


def _line(record: dict[str, Any]) -> dict[str, Any]:
    """从 GeoJSON geometry 或 coordinates_json 构造 LineString。"""

    if record.get("geometry"):
        return record["geometry"]
    return {"type": "LineString", "coordinates": _json_value(record.get("coordinates_json"), "coordinates_json")}


def _build_payload(session: Session, resource: ResourceName, dataset_version_id: int, record: dict[str, Any]) -> Any:
    """把一行导入记录转换为对应的强类型新增请求。"""

    if resource == "rivers":
        return RiverCreate(
            dataset_version_id=dataset_version_id,
            name=str(record.get("name") or "").strip(),
            code=str(record.get("code") or "").strip(),
            length=_float(record.get("length"), "length"),
            level=str(record.get("level") or "main").strip(),
            status=str(record.get("status") or "active").strip(),
            description=record.get("description"),
            geometry=_line(record),
        )
    river_id = _river_id(session, dataset_version_id, record.get("river_code"))
    if resource == "cross_sections":
        points = _json_value(record.get("points_json") or record.get("points"), "points_json")
        if isinstance(points, list):
            points = {"points": points}
        elevation_min = record.get("elevation_min")
        if elevation_min in (None, ""):
            elevation_min = min(float(point[1]) for point in points["points"])
        return CrossSectionCreate(
            dataset_version_id=dataset_version_id,
            river_id=river_id,
            section_code=str(record.get("section_code") or "").strip(),
            section_name=str(record.get("section_name") or "").strip(),
            station=_float(record.get("station"), "station"),
            points=points,
            roughness=_float(record.get("roughness"), "roughness"),
            elevation_min=_float(elevation_min, "elevation_min"),
            survey_date=record.get("survey_date") or None,
            geometry=_point(record),
        )
    if resource == "gates":
        return GateCreate(
            dataset_version_id=dataset_version_id,
            river_id=river_id,
            name=str(record.get("name") or "").strip(),
            gate_code=str(record.get("gate_code") or "").strip(),
            gate_type=str(record.get("gate_type") or "").strip(),
            opening_direction=str(record.get("opening_direction") or "vertical").strip(),
            control_mode=str(record.get("control_mode") or "local").strip(),
            width=_float(record.get("width"), "width"),
            height=_float(record.get("height"), "height"),
            max_flow=_float(record.get("max_flow"), "max_flow"),
            bottom_elevation=_float(record.get("bottom_elevation"), "bottom_elevation"),
            status=str(record.get("status") or "offline").strip(),
            geometry=_point(record),
        )
    curve = _json_value(record.get("efficiency_curve_json") or record.get("efficiency_curve"), "efficiency_curve_json")
    if isinstance(curve, list):
        curve = {"points": curve}
    return PumpCreate(
        dataset_version_id=dataset_version_id,
        river_id=river_id,
        name=str(record.get("name") or "").strip(),
        pump_code=str(record.get("pump_code") or "").strip(),
        design_flow=_float(record.get("design_flow"), "design_flow"),
        head=_float(record.get("head"), "head"),
        power=_float(record.get("power"), "power"),
        efficiency_curve=curve,
        control_mode=str(record.get("control_mode") or "local").strip(),
        status=str(record.get("status") or "offline").strip(),
        geometry=_point(record),
    )


def import_records(session: Session, resource: ResourceName, dataset_version_id: int, records: list[dict[str, Any]], stored_filename: str) -> ImportResponse:
    """先校验全部记录，再在一个事务中写入；任何错误都不产生部分数据。"""

    assert_dataset_version_mutable(session, dataset_version_id)
    payloads: list[Any] = []
    errors: list[ImportIssue] = []
    for row_number, record in enumerate(records, start=2):
        try:
            payloads.append(_build_payload(session, resource, dataset_version_id, record))
        except (ValueError, TypeError, ValidationError, json.JSONDecodeError) as exc:
            errors.append(ImportIssue(row=row_number, message=str(exc)))
    if errors:
        session.rollback()
        return ImportResponse(status="failed", resource=resource, imported_count=0, stored_filename=stored_filename, errors=errors, warnings=[])
    try:
        for payload in payloads:
            if resource == "rivers":
                river_service.create_river(session, payload)
            elif resource == "cross_sections":
                cross_section_service.create_cross_section(session, payload)
            elif resource == "gates":
                structure_service.create_gate(session, payload)
            else:
                structure_service.create_pump(session, payload)
        session.commit()
    except IntegrityError as exc:
        session.rollback()
        return ImportResponse(status="failed", resource=resource, imported_count=0, stored_filename=stored_filename, errors=[ImportIssue(row=1, message="批次违反唯一性、外键或数值约束")], warnings=[])
    return ImportResponse(status="success", resource=resource, imported_count=len(payloads), stored_filename=stored_filename, errors=[], warnings=[])
