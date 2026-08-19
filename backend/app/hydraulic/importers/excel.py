"""Parse bilingual hydraulic network and cross-section Excel templates."""

from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.hydraulic.importers.common import file_identity, safe_code
from app.hydraulic.importers.security import (
    DEFAULT_IMPORT_BUDGET,
    HydraulicImportBudget,
    validate_import_envelope,
    validate_xlsx_cell_budget,
)
from app.hydraulic.schemas import (
    HydraulicBranchInput,
    HydraulicChainageInput,
    HydraulicCrossSectionInput,
    HydraulicExchangePayload,
    HydraulicSectionPointInput,
)


HEADER_ALIASES = {
    "network_code": {"network_code", "网络编码"},
    "network_name": {"network_name", "网络名称"},
    "river_name": {"river_name", "河流名称"},
    "branch_name": {"branch_name", "河段名称"},
    "branch_code": {"branch_code", "河段编码"},
    "flow_direction": {"flow_direction", "流向"},
    "source_revision": {"source_revision", "来源修订"},
    "chainage": {"chainage", "桩号"},
    "x": {"x", "东坐标", "x坐标"},
    "y": {"y", "北坐标", "y坐标"},
    "z": {"z", "高程坐标", "z坐标"},
    "point_code": {"point_code", "点编码"},
    "section_code": {"section_code", "断面编号"},
    "section_name": {"section_name", "断面名称"},
    "topography_id": {"topography_id", "topo_id", "地形编号"},
    "sequence": {"sequence", "点序"},
    "distance": {"distance", "距离"},
    "elevation": {"elevation", "高程"},
    "point_x": {"point_x", "点x", "点东坐标"},
    "point_y": {"point_y", "点y", "点北坐标"},
    "point_z": {"point_z", "点z"},
    "location_x": {"location_x", "位置x", "断面位置x"},
    "location_y": {"location_y", "位置y", "断面位置y"},
    "axis_x": {"axis_x", "断面线x", "轴线x"},
    "axis_y": {"axis_y", "断面线y", "轴线y"},
    "survey_date": {"survey_date", "测量日期"},
    "survey_method": {"survey_method", "测量方法"},
    "default_manning_n": {"default_manning_n", "默认曼宁系数"},
    "marker_type": {"marker_type", "标志类型"},
    "roughness_zone_order": {"roughness_zone_order", "糙率分区序号"},
    "roughness_start": {"roughness_start", "糙率起点"},
    "roughness_end": {"roughness_end", "糙率终点"},
    "roughness_n": {"roughness_n", "曼宁系数"},
    "roughness_type": {"roughness_type", "糙率分区类型"},
}


def _normalized_header(value: object) -> str:
    """Normalize one workbook header for deterministic alias matching."""

    return str(value or "").strip().lower().replace(" ", "").replace("-", "_")


def _header_map(headers: tuple[object, ...]) -> dict[str, int]:
    """Map canonical field names to worksheet column indexes."""

    normalized = [_normalized_header(value) for value in headers]
    mapping: dict[str, int] = {}
    for canonical, aliases in HEADER_ALIASES.items():
        normalized_aliases = {_normalized_header(alias) for alias in aliases}
        for index, header in enumerate(normalized):
            if header in normalized_aliases:
                mapping[canonical] = index
                break
    return mapping


def _rows(sheet: Any) -> list[dict[str, object]]:
    """Return rows below the best canonical header within the first ten lines."""

    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []
    candidates = [(_header_map(row), index) for index, row in enumerate(values[:10])]
    mapping, header_index = max(candidates, key=lambda item: len(item[0]))
    if len(mapping) < 3:
        return []
    rows: list[dict[str, object]] = []
    for raw in values[header_index + 1 :]:
        row = {
            canonical: raw[index]
            for canonical, index in mapping.items()
            if index < len(raw) and raw[index] not in (None, "")
        }
        if row:
            rows.append(row)
    return rows


def _required(row: dict[str, object], key: str, row_number: int) -> object:
    """Read a required cell with an actionable row-level error."""

    if key not in row:
        raise ValueError(f"Excel row {row_number} is missing required field {key}")
    return row[key]


def _as_date(value: object | None) -> date | None:
    """Normalize Excel date or ISO text into a date."""

    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value).strip())


def _is_network_sheet(title: str, rows: list[dict[str, object]]) -> bool:
    """Identify a branch/chainage sheet from its title or required columns."""

    lowered = title.lower()
    return "network" in lowered or "河网" in title or bool(rows and {"branch_code", "x", "y"} <= rows[0].keys())


def _is_section_sheet(title: str, rows: list[dict[str, object]]) -> bool:
    """Identify a cross-section profile sheet from title or point columns."""

    lowered = title.lower()
    return "section" in lowered or "断面" in title or bool(rows and {"section_code", "distance", "elevation"} <= rows[0].keys())


def parse_excel(
    filename: str,
    content: bytes,
    source_srid: int,
    *,
    budget: HydraulicImportBudget = DEFAULT_IMPORT_BUDGET,
) -> HydraulicExchangePayload:
    """Parse one or both official hydraulic workbook layouts into a common payload."""

    validate_import_envelope(filename, content, budget=budget)
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    validate_xlsx_cell_budget(workbook, budget=budget)
    default_code, default_name = file_identity(filename, "HYDRAULIC-XLSX")
    network_code = default_code
    network_name = default_name
    branch_rows: list[dict[str, object]] = []
    section_rows: list[dict[str, object]] = []
    for sheet in workbook.worksheets:
        rows = _rows(sheet)
        if not rows:
            continue
        first = rows[0]
        network_code = safe_code(str(first.get("network_code", network_code)), network_code)
        network_name = str(first.get("network_name", network_name))[:128]
        if _is_network_sheet(sheet.title, rows):
            branch_rows.extend(rows)
        if _is_section_sheet(sheet.title, rows):
            section_rows.extend(rows)

    grouped_branches: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row_number, row in enumerate(branch_rows, start=2):
        code = safe_code(str(_required(row, "branch_code", row_number)), f"BRANCH-{row_number:03d}")
        grouped_branches[code].append(row)
    branches: list[HydraulicBranchInput] = []
    for code, rows in grouped_branches.items():
        ordered = sorted(rows, key=lambda row: float(row.get("chainage", 0)))
        first = ordered[0]
        branches.append(
            HydraulicBranchInput(
                code=code,
                river_name=str(first.get("river_name") or first.get("branch_name") or code)[:128],
                branch_name=str(first.get("branch_name") or code)[:128],
                flow_direction=str(first.get("flow_direction", "forward")).strip().lower(),
                source_revision=str(first["source_revision"])[:64] if first.get("source_revision") else None,
                points=[
                    HydraulicChainageInput(
                        chainage=float(_required(row, "chainage", index)),
                        x=float(_required(row, "x", index)),
                        y=float(_required(row, "y", index)),
                        z=float(row["z"]) if "z" in row else None,
                        point_code=str(row["point_code"])[:64] if row.get("point_code") else None,
                    )
                    for index, row in enumerate(ordered, start=2)
                ],
            )
        )

    grouped_sections: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row_number, row in enumerate(section_rows, start=2):
        code = safe_code(str(_required(row, "section_code", row_number)), f"XS-{row_number:04d}")
        grouped_sections[code].append(row)
    sections: list[HydraulicCrossSectionInput] = []
    for code, rows in grouped_sections.items():
        ordered = sorted(rows, key=lambda row: int(row.get("sequence", 0)))
        first = ordered[0]
        location_x = first.get("location_x")
        location_y = first.get("location_y")
        axis_points = [
            (float(row["axis_x"]), float(row["axis_y"]))
            for row in ordered
            if "axis_x" in row and "axis_y" in row
        ]
        sections.append(
            HydraulicCrossSectionInput(
                section_code=code,
                section_name=str(first["section_name"])[:128] if first.get("section_name") else None,
                branch_code=safe_code(
                    str(_required(first, "branch_code", 2)), "BRANCH-UNKNOWN"
                ),
                chainage=float(_required(first, "chainage", 2)),
                topography_id=str(first.get("topography_id", "DEFAULT"))[:64],
                survey_date=_as_date(first.get("survey_date")),
                survey_method=str(first["survey_method"])[:64] if first.get("survey_method") else None,
                default_manning_n=float(first.get("default_manning_n", 0.03)),
                location_x=float(location_x) if location_x is not None else None,
                location_y=float(location_y) if location_y is not None else None,
                axis_points=axis_points,
                roughness_zones=[
                    {
                        "zone_order": int(row["roughness_zone_order"]),
                        "offset_start_m": float(row["roughness_start"]),
                        "offset_end_m": float(row["roughness_end"]),
                        "manning_n": float(row["roughness_n"]),
                        "zone_type": str(row.get("roughness_type") or "channel")[:32],
                    }
                    for row in ordered
                    if all(key in row for key in (
                        "roughness_zone_order", "roughness_start", "roughness_end", "roughness_n"
                    ))
                ],
                points=[
                    HydraulicSectionPointInput(
                        sequence=int(row.get("sequence", index)),
                        distance=float(_required(row, "distance", index + 2)),
                        elevation=float(_required(row, "elevation", index + 2)),
                        marker_type=str(row.get("marker_type") or "none").lower(),
                        point_code=str(row["point_code"])[:64] if row.get("point_code") else None,
                        x=float(row["point_x"]) if "point_x" in row else None,
                        y=float(row["point_y"]) if "point_y" in row else None,
                        z=float(row["point_z"]) if "point_z" in row else None,
                    )
                    for index, row in enumerate(ordered)
                ],
            )
        )
    return HydraulicExchangePayload(
        network_code=network_code,
        network_name=network_name,
        source_srid=source_srid,
        source_kind="excel",
        branches=branches,
        sections=sections,
    )
