"""Structured engineering result products and safe CSV/XLSX/GeoJSON exports."""

from __future__ import annotations

import csv
from io import BytesIO, StringIO
import json
from typing import Any, Iterable

from openpyxl import Workbook

from app.hydraulic.production.contracts import (
    HydraulicResultPoint,
    ResultProductBundle,
    ResultProductRequest,
)


def _group_points(
    points: Iterable[HydraulicResultPoint],
) -> dict[tuple[str, str, str, float], list[HydraulicResultPoint]]:
    """Group by immutable scenario/location identity."""

    grouped: dict[tuple[str, str, str, float], list[HydraulicResultPoint]] = {}
    for point in points:
        key = (
            point.scenario_id,
            point.branch_id,
            point.cross_section_id,
            float(point.chainage_m),
        )
        grouped.setdefault(key, []).append(point)
    return grouped


def _single_optional(values: Iterable[float | None], field: str) -> float | None:
    """Reject conflicting source attributes rather than choosing one silently."""

    available = {float(value) for value in values if value is not None}
    if len(available) > 1:
        raise ValueError(f"result points contain conflicting {field}")
    return next(iter(available)) if available else None


def _envelope(points: list[HydraulicResultPoint]) -> dict[str, Any]:
    """Build one location envelope with independent maxima and peak times."""

    first = points[0]
    maximum_water = max(points, key=lambda item: item.water_level_m)
    maximum_discharge = max(points, key=lambda item: item.discharge_m3s)
    maximum_velocity = max(points, key=lambda item: item.velocity_m_s)
    maximum_depth = max(
        (item for item in points if item.depth_m is not None),
        key=lambda item: float(item.depth_m or 0.0),
        default=None,
    )
    geometry_values = [
        json.dumps(item.geometry, sort_keys=True, separators=(",", ":"))
        for item in points
        if item.geometry is not None
    ]
    if len(set(geometry_values)) > 1:
        raise ValueError("result points contain conflicting location geometry")
    return {
        "scenario_id": first.scenario_id,
        "branch_id": first.branch_id,
        "cross_section_id": first.cross_section_id,
        "chainage_m": float(first.chainage_m),
        "maximum_water_level_m": float(maximum_water.water_level_m),
        "water_level_peak_time_seconds": float(maximum_water.time_seconds),
        "maximum_depth_m": float(maximum_depth.depth_m) if maximum_depth else None,
        "depth_peak_time_seconds": float(maximum_depth.time_seconds) if maximum_depth else None,
        "maximum_discharge_m3s": float(maximum_discharge.discharge_m3s),
        "discharge_peak_time_seconds": float(maximum_discharge.time_seconds),
        "maximum_velocity_m_s": float(maximum_velocity.velocity_m_s),
        "velocity_peak_time_seconds": float(maximum_velocity.time_seconds),
        "bed_elevation_m": _single_optional(
            (item.bed_elevation_m for item in points), "bed_elevation_m"
        ),
        "left_bank_elevation_m": _single_optional(
            (item.left_bank_elevation_m for item in points), "left_bank_elevation_m"
        ),
        "right_bank_elevation_m": _single_optional(
            (item.right_bank_elevation_m for item in points), "right_bank_elevation_m"
        ),
        "geometry": json.loads(geometry_values[0]) if geometry_values else None,
    }


def _scenario_differences(
    request: ResultProductRequest,
) -> list[dict[str, Any]]:
    """Calculate exact-time project-minus-baseline water-level differences."""

    if request.baseline_scenario_id is None:
        return []
    baseline = {
        (item.branch_id, item.cross_section_id, float(item.chainage_m), float(item.time_seconds)): item
        for item in request.points
        if item.scenario_id == request.baseline_scenario_id
    }
    project = {
        (item.branch_id, item.cross_section_id, float(item.chainage_m), float(item.time_seconds)): item
        for item in request.points
        if item.scenario_id == request.project_scenario_id
    }
    shared = sorted(set(baseline) & set(project), key=lambda item: (item[0], item[2], item[3]))
    if not shared:
        raise ValueError("baseline and project scenarios have no exact shared result identities")
    return [
        {
            "branch_id": key[0],
            "cross_section_id": key[1],
            "chainage_m": key[2],
            "time_seconds": key[3],
            "baseline_water_level_m": float(baseline[key].water_level_m),
            "project_water_level_m": float(project[key].water_level_m),
            "delta_h_m": float(project[key].water_level_m - baseline[key].water_level_m),
        }
        for key in shared
    ]


def _afflux_reaches(
    differences: list[dict[str, Any]], threshold: float
) -> tuple[dict[str, Any] | None, list[dict[str, Any]]]:
    """Find maximum afflux and consecutive above-threshold one-dimensional reaches."""

    if not differences:
        return None, []
    maximum = max(differences, key=lambda item: float(item["delta_h_m"]))
    by_location: dict[tuple[str, str, float], dict[str, Any]] = {}
    for item in differences:
        key = (str(item["branch_id"]), str(item["cross_section_id"]), float(item["chainage_m"]))
        current = by_location.get(key)
        if current is None or float(item["delta_h_m"]) > float(current["delta_h_m"]):
            by_location[key] = item
    reaches: list[dict[str, Any]] = []
    branches = sorted({key[0] for key in by_location})
    for branch_id in branches:
        ordered = sorted(
            (item for key, item in by_location.items() if key[0] == branch_id),
            key=lambda item: float(item["chainage_m"]),
        )
        active: list[dict[str, Any]] = []
        for item in ordered + [None]:
            if item is not None and float(item["delta_h_m"]) >= threshold:
                active.append(item)
                continue
            if active:
                reaches.append(
                    {
                        "branch_id": branch_id,
                        "start_chainage_m": float(active[0]["chainage_m"]),
                        "end_chainage_m": float(active[-1]["chainage_m"]),
                        "maximum_afflux_m": max(float(value["delta_h_m"]) for value in active),
                        "threshold_m": threshold,
                        "section_count": len(active),
                    }
                )
                active = []
    maximum_afflux = {
        "maximum_afflux_m": float(maximum["delta_h_m"]),
        "branch_id": maximum["branch_id"],
        "cross_section_id": maximum["cross_section_id"],
        "chainage_m": maximum["chainage_m"],
        "time_seconds": maximum["time_seconds"],
        "threshold_m": threshold,
    }
    return maximum_afflux, reaches


def build_result_products(request: ResultProductRequest) -> ResultProductBundle:
    """Generate envelopes, profiles, deltas, afflux, tables, and GeoJSON."""

    grouped = _group_points(request.points)
    envelope = [
        _envelope(grouped[key]) for key in sorted(grouped, key=lambda item: (item[0], item[1], item[3]))
    ]
    project_envelope = [
        item for item in envelope if item["scenario_id"] == request.project_scenario_id
    ]
    if not project_envelope:
        raise ValueError("project scenario has no result points")
    profile = [
        {
            "branch_id": item["branch_id"],
            "cross_section_id": item["cross_section_id"],
            "chainage_m": item["chainage_m"],
            "bed_elevation_m": item["bed_elevation_m"],
            "design_water_level_m": item["maximum_water_level_m"],
            "left_bank_elevation_m": item["left_bank_elevation_m"],
            "right_bank_elevation_m": item["right_bank_elevation_m"],
        }
        for item in project_envelope
    ]
    differences = _scenario_differences(request)
    maximum_afflux, reaches = _afflux_reaches(
        differences, float(request.afflux_threshold_m)
    )
    baseline_by_location = {
        (item["branch_id"], item["cross_section_id"], item["chainage_m"]): item
        for item in envelope
        if request.baseline_scenario_id is not None
        and item["scenario_id"] == request.baseline_scenario_id
    }
    max_delta_by_location: dict[tuple[str, str, float], float] = {}
    for item in differences:
        key = (str(item["branch_id"]), str(item["cross_section_id"]), float(item["chainage_m"]))
        max_delta_by_location[key] = max(
            max_delta_by_location.get(key, float("-inf")), float(item["delta_h_m"])
        )
    key_sections: list[dict[str, Any]] = []
    for item in project_envelope:
        key = (str(item["branch_id"]), str(item["cross_section_id"]), float(item["chainage_m"]))
        baseline = baseline_by_location.get(key)
        key_sections.append(
            {
                "cross_section_id": item["cross_section_id"],
                "branch_id": item["branch_id"],
                "chainage_m": item["chainage_m"],
                "bed_elevation_m": item["bed_elevation_m"],
                "baseline_hmax_m": baseline["maximum_water_level_m"] if baseline else None,
                "project_hmax_m": item["maximum_water_level_m"],
                "delta_h_m": max_delta_by_location.get(key),
                "qmax_m3s": item["maximum_discharge_m3s"],
                "vmax_m_s": item["maximum_velocity_m_s"],
                "peak_time_seconds": item["water_level_peak_time_seconds"],
            }
        )
    features = [
        {
            "type": "Feature",
            "geometry": item["geometry"],
            "properties": {key: value for key, value in item.items() if key != "geometry"},
        }
        for item in project_envelope
    ]
    return ResultProductBundle(
        max_envelope=envelope,
        longitudinal_profile=profile,
        scenario_difference=differences,
        maximum_afflux=maximum_afflux,
        afflux_reaches=reaches,
        key_section_table=key_sections,
        calibration_table=request.calibration_table,
        validation_table=request.validation_table,
        external_comparison_table=request.external_comparison_table,
        geojson={
            "type": "FeatureCollection",
            "metadata": {
                "project_id": request.project_id,
                "model_version": request.model_version,
                "scenario_id": request.project_scenario_id,
                "geometry_null_means_unavailable": True,
            },
            "features": features,
        },
    )


def _safe_cell(value: object) -> object:
    """Neutralize spreadsheet formula injection while preserving displayed text."""

    if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        return "'" + value
    return value


def _tabular_rows(bundle: ResultProductBundle) -> dict[str, list[dict[str, Any]]]:
    """Return dynamic workbook tables from the common product bundle."""

    tables = {
        "Max Results": bundle.max_envelope,
        "Longitudinal Profile": bundle.longitudinal_profile,
        "Scenario Compare": bundle.scenario_difference,
        "Afflux Reaches": bundle.afflux_reaches,
        "Key Sections": bundle.key_section_table,
        "Calibration": bundle.calibration_table,
        "Validation": bundle.validation_table,
        "External Compare": bundle.external_comparison_table,
    }
    return {name: rows for name, rows in tables.items() if rows}


def export_product_csv(bundle: ResultProductBundle, table: str = "key_section_table") -> bytes:
    """Export one named product table as UTF-8 CSV."""

    rows = getattr(bundle, table, None)
    if not isinstance(rows, list) or not rows:
        raise ValueError(f"product table {table!r} is unavailable or empty")
    headers = sorted({key for row in rows for key in row if key != "geometry"})
    output = StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(
        {key: _safe_cell(value) for key, value in row.items() if key != "geometry"}
        for row in rows
    )
    return output.getvalue().encode("utf-8-sig")


def export_product_xlsx(bundle: ResultProductBundle) -> bytes:
    """Export available engineering tables with safe user-originated strings."""

    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    summary = workbook.create_sheet("Summary")
    summary.append(["Product", "Rows"])
    tables = _tabular_rows(bundle)
    for name, rows in tables.items():
        summary.append([_safe_cell(name), len(rows)])
        sheet = workbook.create_sheet(name[:31])
        headers = sorted({key for row in rows for key in row if key != "geometry"})
        sheet.append(headers)
        for row in rows:
            sheet.append([_safe_cell(row.get(header)) for header in headers])
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def export_product_geojson(bundle: ResultProductBundle) -> bytes:
    """Export the factual result feature collection as UTF-8 GeoJSON."""

    return json.dumps(bundle.geojson, ensure_ascii=False, separators=(",", ":")).encode("utf-8")


__all__ = [
    "build_result_products",
    "export_product_csv",
    "export_product_geojson",
    "export_product_xlsx",
]
