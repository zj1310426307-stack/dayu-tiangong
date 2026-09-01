"""Dry-run engineering and external-result import facade with explicit mappings."""

from __future__ import annotations

import csv
from datetime import UTC, datetime
from hashlib import sha256
from io import BytesIO, StringIO
from pathlib import Path
from typing import Mapping
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import load_workbook

from app.hydraulic.importers import parse_hydraulic_file
from app.hydraulic.importers.security import DEFAULT_IMPORT_BUDGET, validate_import_envelope
from app.hydraulic.production.contracts import (
    ExternalResultImportOptions,
    ExternalResultPoint,
    ExternalResultPreview,
    ProductionSeries,
    ProductionSeriesPoint,
    QAIssue,
    TimeSeriesImportOptions,
    TimeSeriesImportPreview,
    finite_number,
)


def _safe_filename(filename: str) -> str:
    """Keep an original basename as metadata, never as a storage path."""

    value = Path(filename).name
    if not value or value in {".", ".."}:
        raise ValueError("source filename is invalid")
    return value[:255]


def _csv_rows(content: bytes) -> list[dict[str, object]]:
    """Read bounded UTF-8 CSV rows while rejecting undeclared columns."""

    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("external CSV must be UTF-8 or UTF-8 with BOM") from exc
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        raise ValueError("external CSV header is required")
    if len(reader.fieldnames) > DEFAULT_IMPORT_BUDGET.max_csv_columns:
        raise ValueError("external CSV exceeds the configured column limit")
    rows: list[dict[str, object]] = []
    for index, row in enumerate(reader, start=1):
        if index > DEFAULT_IMPORT_BUDGET.max_csv_rows:
            raise ValueError("external CSV exceeds the configured row limit")
        if None in row:
            raise ValueError(f"external CSV row {index + 1} contains undeclared columns")
        rows.append({str(key).strip(): value for key, value in row.items()})
    return rows


def _xlsx_rows(content: bytes, sheet_name: str | None) -> list[dict[str, object]]:
    """Read one bounded worksheet and reject formulas in mapped source data."""

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"external workbook has no sheet named {sheet_name!r}")
            sheet = workbook[sheet_name]
        else:
            sheet = workbook[workbook.sheetnames[0]]
        iterator = sheet.iter_rows(values_only=True)
        try:
            header_values = next(iterator)
        except StopIteration as exc:
            raise ValueError("external workbook sheet is empty") from exc
        headers = [str(value).strip() if value is not None else "" for value in header_values]
        if not headers or any(not value for value in headers) or len(headers) != len(set(headers)):
            raise ValueError("external workbook headers must be non-empty and unique")
        if len(headers) > DEFAULT_IMPORT_BUDGET.max_csv_columns:
            raise ValueError("external workbook exceeds the configured column limit")
        rows: list[dict[str, object]] = []
        for index, values in enumerate(iterator, start=1):
            if index > DEFAULT_IMPORT_BUDGET.max_csv_rows:
                raise ValueError("external workbook exceeds the configured row limit")
            row = dict(zip(headers, values))
            if all(value is None for value in row.values()):
                continue
            formulas = [key for key, value in row.items() if isinstance(value, str) and value.startswith("=")]
            if formulas:
                raise ValueError(
                    f"external workbook row {index + 1} contains formulas in: {', '.join(formulas)}"
                )
            rows.append(row)
        return rows
    finally:
        workbook.close()


def _absolute_time(value: object, timezone_name: str | None) -> datetime:
    """Parse one absolute time with an explicit and stable timezone policy."""

    if isinstance(value, datetime):
        parsed = value
    else:
        text = str(value).strip()
        if text.endswith("Z"):
            text = text[:-1] + "+00:00"
        try:
            parsed = datetime.fromisoformat(text)
        except ValueError as exc:
            raise ValueError(f"invalid ISO-8601 external time {value!r}") from exc
    if parsed.tzinfo is None:
        if not timezone_name:
            raise ValueError("naive external timestamps require an explicit timezone")
        try:
            parsed = parsed.replace(tzinfo=ZoneInfo(timezone_name))
        except ZoneInfoNotFoundError as exc:
            raise ValueError(f"unknown timezone {timezone_name!r}") from exc
    return parsed.astimezone(UTC)


def _mapped_value(row: Mapping[str, object], column: str, row_number: int) -> object:
    """Read an explicitly mapped column with a row-specific error."""

    if column not in row:
        raise ValueError(f"mapped column {column!r} does not exist")
    value = row[column]
    if value is None or str(value).strip() == "":
        raise ValueError(f"row {row_number} mapped column {column!r} is empty")
    return value


class EngineeringDataImporter:
    """Inspect and preview engineering data without writing authoritative rows."""

    def inspect(self, filename: str, content: bytes) -> dict[str, object]:
        """Return safe identity metadata and enforce the shared upload envelope."""

        safe_name = _safe_filename(filename)
        validate_import_envelope(safe_name, content)
        return {
            "original_filename": safe_name,
            "extension": Path(safe_name).suffix.lower(),
            "size_bytes": len(content),
            "sha256": sha256(content).hexdigest(),
        }

    def preview_hydraulic(
        self, filename: str, content: bytes, source_srid: int
    ) -> dict[str, object]:
        """Reuse the existing normalized network/section pipeline for a dry run."""

        identity = self.inspect(filename, content)
        payload, parser_profile, native_status = parse_hydraulic_file(
            str(identity["original_filename"]), content, source_srid
        )
        issues: list[dict[str, object]] = []
        if not payload.branches:
            issues.append(
                {
                    "code": "IMPORT_NETWORK_BRANCH_MISSING",
                    "severity": "ERROR",
                    "message": "Import contains no Branch records.",
                }
            )
        if payload.coordinate_reference is None:
            issues.append(
                {
                    "code": "IMPORT_CRS_CONTRACT_REQUIRED",
                    "severity": "ERROR",
                    "message": "Import requires a complete coordinate reference contract.",
                }
            )
        return {
            **identity,
            "parser_profile": parser_profile,
            "native_status": native_status,
            "counts": {
                "branches": len(payload.branches),
                "cross_sections": len(payload.sections),
                "cross_section_points": sum(len(item.points) for item in payload.sections),
            },
            "issues": issues,
            "payload": payload.model_dump(mode="json"),
            "commit_allowed": not any(item["severity"] == "ERROR" for item in issues),
        }

    def preview_external(
        self,
        filename: str,
        content: bytes,
        options: ExternalResultImportOptions,
    ) -> ExternalResultPreview:
        """Normalize legal CSV/XLSX exports with explicit column and location mapping."""

        identity = self.inspect(filename, content)
        suffix = str(identity["extension"])
        if suffix == ".csv":
            rows = _csv_rows(content)
        elif suffix == ".xlsx":
            rows = _xlsx_rows(content, options.sheet_name)
        else:
            raise ValueError("external hydraulic results support .csv or .xlsx only")
        if not rows:
            raise ValueError("external result contains no data rows")
        branch_mappings = {item.external_branch: item for item in options.branch_mappings}
        if len(branch_mappings) != len(options.branch_mappings):
            raise ValueError("external branch mappings must be unique")
        mapping = options.column_mapping
        parsed_times: list[float | datetime] = []
        for index, row in enumerate(rows, start=2):
            raw_time = _mapped_value(row, mapping.time, index)
            parsed_times.append(
                finite_number(raw_time, mapping.time)
                if options.time_basis == "relative"
                else _absolute_time(raw_time, options.timezone)
            )
        origin = (
            min(value for value in parsed_times if isinstance(value, datetime))
            if options.time_basis == "absolute"
            else None
        )
        points: list[ExternalResultPoint] = []
        for row_index, (row, parsed_time) in enumerate(zip(rows, parsed_times), start=2):
            external_branch = str(_mapped_value(row, mapping.branch, row_index)).strip()
            branch_mapping = branch_mappings.get(external_branch)
            if branch_mapping is None:
                raise ValueError(
                    f"row {row_index} external Branch {external_branch!r} has no confirmed mapping"
                )
            external_chainage = finite_number(
                _mapped_value(row, mapping.chainage, row_index), mapping.chainage
            )
            chainage = branch_mapping.map_chainage(external_chainage)
            if chainage < 0:
                raise ValueError(f"row {row_index} maps to negative Dayu chainage")
            timestamp = parsed_time if isinstance(parsed_time, datetime) else None
            time_seconds = (
                (timestamp - origin).total_seconds()
                if timestamp is not None and origin is not None
                else float(parsed_time)
            )

            def optional(column: str | None) -> float | None:
                if not column:
                    return None
                return finite_number(_mapped_value(row, column, row_index), column)

            points.append(
                ExternalResultPoint(
                    external_branch=external_branch,
                    branch_id=branch_mapping.dayu_branch,
                    external_chainage=external_chainage,
                    chainage_m=chainage,
                    time_seconds=time_seconds,
                    timestamp=timestamp,
                    water_level_m=optional(mapping.water_level),
                    discharge_m3s=optional(mapping.discharge),
                    velocity_m_s=optional(mapping.velocity),
                )
            )
        identities = [
            (point.branch_id, point.chainage_m, point.time_seconds) for point in points
        ]
        if len(identities) != len(set(identities)):
            raise ValueError("external result contains duplicate Branch/chainage/time rows")
        variables = [
            variable
            for variable, column in (
                ("water_level", mapping.water_level),
                ("discharge", mapping.discharge),
                ("velocity", mapping.velocity),
            )
            if column
        ]
        issues: list[QAIssue] = []
        if options.external_model_version == "UNKNOWN":
            issues.append(
                QAIssue(
                    code="EXTERNAL_MODEL_VERSION_UNKNOWN",
                    severity="INFO",
                    category="Observation",
                    entity_type="ExternalHydraulicResult",
                    message="External model version was not supplied and remains UNKNOWN.",
                )
            )
        if options.vertical_datum.upper() == "UNKNOWN" and mapping.water_level:
            issues.append(
                QAIssue(
                    code="EXTERNAL_RESULT_DATUM_UNKNOWN",
                    severity="WARNING",
                    category="CRS",
                    entity_type="ExternalHydraulicResult",
                    message="External water levels have an unknown vertical datum.",
                )
            )
        return ExternalResultPreview(
            source_filename=str(identity["original_filename"]),
            source_sha256=str(identity["sha256"]),
            row_count=len(points),
            branch_count=len({point.branch_id for point in points}),
            variables=variables,
            issues=issues,
            points=points,
            provenance={
                "external_model_name": options.external_model_name,
                "external_model_version": options.external_model_version,
                "scenario": options.scenario,
                "vertical_datum": options.vertical_datum,
                "time_basis": options.time_basis,
                "timezone": options.timezone,
                "column_mapping": mapping.model_dump(mode="json"),
                "branch_mappings": [item.model_dump(mode="json") for item in options.branch_mappings],
                "imported_at": datetime.now(UTC).isoformat(),
            },
        )

    def preview_series(
        self,
        filename: str,
        content: bytes,
        options: TimeSeriesImportOptions,
    ) -> TimeSeriesImportPreview:
        """Normalize a boundary or observation CSV/XLSX as a dry run."""

        identity = self.inspect(filename, content)
        suffix = str(identity["extension"])
        if suffix == ".csv":
            rows = _csv_rows(content)
        elif suffix == ".xlsx":
            rows = _xlsx_rows(content, options.sheet_name)
        else:
            raise ValueError("boundary/observation series support .csv or .xlsx only")
        if not rows:
            raise ValueError("time-series import contains no data rows")
        mapping = options.column_mapping
        parsed_times: list[float | datetime] = []
        for index, row in enumerate(rows, start=2):
            raw_time = _mapped_value(row, mapping.time, index)
            parsed_times.append(
                finite_number(raw_time, mapping.time)
                if options.time_basis == "relative"
                else _absolute_time(raw_time, options.timezone)
            )
        origin = (
            min(value for value in parsed_times if isinstance(value, datetime))
            if options.time_basis == "absolute"
            else None
        )
        samples: list[ProductionSeriesPoint] = []
        for index, (row, parsed_time) in enumerate(zip(rows, parsed_times), start=2):
            raw_value = row.get(mapping.value)
            quality = str(row.get(mapping.quality_flag, "GOOD") or "GOOD").strip().upper()
            if quality not in {"GOOD", "SUSPECT", "MISSING", "REJECTED"}:
                raise ValueError(f"row {index} has unsupported quality flag {quality!r}")
            missing = quality in {"MISSING", "REJECTED"} or raw_value in (None, "")
            if missing and quality in {"GOOD", "SUSPECT"}:
                quality = "MISSING"
            timestamp = parsed_time if isinstance(parsed_time, datetime) else None
            time_seconds = (
                (timestamp - origin).total_seconds()
                if timestamp is not None and origin is not None
                else float(parsed_time)
            )
            samples.append(
                ProductionSeriesPoint(
                    time_seconds=time_seconds,
                    timestamp=timestamp,
                    value=None if missing else finite_number(raw_value, mapping.value),
                    quality_flag=quality,
                )
            )
        series = ProductionSeries(
            series_id=options.series_id,
            variable=options.variable,
            unit=options.unit,
            samples=samples,
            source=options.source,
            branch_id=options.branch_id,
            chainage_m=options.chainage_m,
            station_id=options.station_id,
            vertical_datum=options.vertical_datum,
            time_basis=options.time_basis,
            timezone=options.timezone,
        )
        issues: list[QAIssue] = []
        missing_count = sum(sample.quality_flag == "MISSING" for sample in samples)
        if missing_count:
            issues.append(
                QAIssue(
                    code="IMPORT_SERIES_MISSING_VALUES",
                    severity="WARNING",
                    category="Observation" if options.series_kind == "observation" else "Boundary",
                    entity_type="ObservationSeries" if options.series_kind == "observation" else "BoundarySeries",
                    entity_id=options.series_id,
                    message="Missing samples remain missing and were not filled with zero.",
                    context={"missing_sample_count": missing_count},
                )
            )
        if options.variable == "water_level" and options.vertical_datum.upper() == "UNKNOWN":
            issues.append(
                QAIssue(
                    code="IMPORT_SERIES_DATUM_UNKNOWN",
                    severity="WARNING",
                    category="CRS",
                    entity_type="ObservationSeries" if options.series_kind == "observation" else "BoundarySeries",
                    entity_id=options.series_id,
                    message="Water-level series vertical datum remains UNKNOWN.",
                )
            )
        return TimeSeriesImportPreview(
            source_filename=str(identity["original_filename"]),
            source_sha256=str(identity["sha256"]),
            row_count=len(samples),
            issues=issues,
            series=series,
            provenance={
                "series_kind": options.series_kind,
                "column_mapping": mapping.model_dump(mode="json"),
                "imported_at": datetime.now(UTC).isoformat(),
            },
        )


__all__ = ["EngineeringDataImporter"]
