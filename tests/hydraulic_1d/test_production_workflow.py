"""HYDRO-1D-PRODUCTION-04 acceptance contracts P01 through P06."""

from __future__ import annotations

from datetime import UTC, datetime, timedelta
from io import BytesIO

from openpyxl import load_workbook
import pytest

from app.hydraulic.production.calibration import (
    build_parameter_sweep,
    evaluate_acceptance,
    evaluate_project_metric_criteria,
    evaluate_validation_independence,
    rank_calibration_candidates,
)
from app.hydraulic.production.comparison import compare_external_result
from app.hydraulic.production.contracts import (
    AcceptanceCriteria,
    AcceptanceEvaluationRequest,
    CalibrationObjective,
    CalibrationParameter,
    CalibrationRankingRequest,
    DatasetWindow,
    ExternalBranchMapping,
    ExternalColumnMapping,
    ExternalComparisonRequest,
    ExternalResultImportOptions,
    HydraulicModelQARequest,
    HydraulicResultPoint,
    ParameterSweepRequest,
    ProductionBoundary,
    ProductionBranch,
    ProductionCrossSection,
    ProductionSeries,
    ProductionSeriesPoint,
    ProductionStructure,
    ResultProductRequest,
    TimeAlignmentOptions,
    TimeSeriesColumnMapping,
    TimeSeriesImportOptions,
)
from app.hydraulic.production.gate import (
    assert_production_gate,
    build_production_gate,
)
from app.hydraulic.production.importers import EngineeringDataImporter
from app.hydraulic.production.metrics import align_and_score
from app.hydraulic.production.products import (
    build_result_products,
    export_product_csv,
    export_product_geojson,
    export_product_xlsx,
)
from app.hydraulic.production.qa import HydraulicModelQA
from model.hydraulic_1d import (
    BoundaryCondition,
    CrossSectionPoint,
    Hydraulic1DModel,
    HydraulicBranch,
    HydraulicCrossSection,
    InitialCondition,
    SimulationSettings,
    TimeValue,
)
from model.hydraulic_1d.errors import Hydraulic1DValidationError
from model.provenance import snapshot_hash
from app.model_engine.schemas import SimulationTaskCreate


def _series(
    series_id: str,
    variable: str,
    values: list[float | None],
    *,
    source: str = "synthetic-contract-fixture",
    branch_id: str = "1",
    chainage_m: float = 0.0,
    datum: str = "1985-national-height",
) -> ProductionSeries:
    unit = {"water_level": "m", "discharge": "m3/s", "velocity": "m/s"}[variable]
    return ProductionSeries(
        series_id=series_id,
        variable=variable,
        unit=unit,
        source=source,
        branch_id=branch_id,
        chainage_m=chainage_m,
        station_id=series_id,
        vertical_datum=datum,
        samples=[
            ProductionSeriesPoint(
                time_seconds=index * 60.0,
                value=value,
                quality_flag="GOOD" if value is not None else "MISSING",
            )
            for index, value in enumerate(values)
        ],
    )


def _valid_qa() -> HydraulicModelQARequest:
    branch = ProductionBranch(
        branch_id="1",
        start_chainage_m=0,
        end_chainage_m=100,
        direction_confirmed=True,
        centerline=[(0, 0), (100, 0)],
        upstream_node_id="N1",
        downstream_node_id="N2",
    )
    sections = [
        ProductionCrossSection(
            section_id=str(index + 1),
            branch_id="1",
            chainage_m=chainage,
            offsets_m=[0, 5, 10],
            elevations_m=[2 - index * 0.1, 0 - index * 0.1, 2 - index * 0.1],
            vertical_datum="1985-national-height",
            orientation_confirmed=True,
            axis=[(chainage, -5), (chainage, 5)],
            location=(chainage, 0),
        )
        for index, chainage in enumerate((0.0, 100.0))
    ]
    boundaries = [
        ProductionBoundary(
            boundary_id="U",
            branch_id="1",
            location="upstream",
            series=_series("U", "discharge", [10, 12, 11]),
        ),
        ProductionBoundary(
            boundary_id="D",
            branch_id="1",
            location="downstream",
            series=_series("D", "water_level", [3, 3.2, 3.1], chainage_m=100),
        ),
    ]
    return HydraulicModelQARequest(
        engineering_crs="EPSG:4547",
        horizontal_unit="m",
        vertical_datum="1985-national-height",
        simulation_duration_seconds=120,
        branches=[branch],
        cross_sections=sections,
        boundaries=boundaries,
    )


def _frozen_model() -> Hydraulic1DModel:
    return Hydraulic1DModel(
        simulation_id="sim-production-test",
        scenario_id="case-1",
        network_id="network-1",
        branches=(
            HydraulicBranch(
                id="1",
                code="B1",
                upstream_node_id="N1",
                downstream_node_id="N2",
                start_chainage_m=0,
                end_chainage_m=100,
            ),
        ),
        cross_sections=tuple(
            HydraulicCrossSection(
                id=str(index + 1),
                branch_id="1",
                code=f"XS-{index + 1}",
                chainage_m=chainage,
                vertical_datum="1985-national-height",
                points=(
                    CrossSectionPoint(station_m=0, elevation_m=2),
                    CrossSectionPoint(station_m=5, elevation_m=0),
                    CrossSectionPoint(station_m=10, elevation_m=2),
                ),
                manning_n=0.03,
            )
            for index, chainage in enumerate((0.0, 100.0))
        ),
        boundaries=(
            BoundaryCondition(
                id="U",
                branch_id="1",
                location="upstream",
                variable="discharge",
                series=(TimeValue(time_seconds=0, value=10),),
            ),
            BoundaryCondition(
                id="D",
                branch_id="1",
                location="downstream",
                variable="water_level",
                series=(TimeValue(time_seconds=0, value=3),),
            ),
        ),
        initial_condition=InitialCondition(water_level_m=3, discharge_m3s=10),
        settings=SimulationSettings(
            duration_seconds=120,
            time_step_seconds=10,
            output_interval_seconds=60,
        ),
        metadata={
            "engineering_crs": "EPSG:4547",
            "vertical_datum": "1985-national-height",
        },
    )


def test_p01_csv_xlsx_series_import_keeps_missing_values_and_lineage() -> None:
    """P01: CSV/XLSX import is explicit, traceable, and never fills missing with zero."""

    options = TimeSeriesImportOptions(
        series_kind="observation",
        series_id="OBS-H-01",
        variable="water_level",
        unit="m",
        source="survey-team-A",
        branch_id="1",
        chainage_m=50,
        station_id="STA-01",
        vertical_datum="1985-national-height",
        time_basis="relative",
        column_mapping=TimeSeriesColumnMapping(
            time="seconds", value="stage", quality_flag="quality"
        ),
    )
    content = b"seconds,stage,quality\n0,3.1,GOOD\n60,,MISSING\n120,3.4,GOOD\n"
    preview = EngineeringDataImporter().preview_series("observed.csv", content, options)
    assert preview.row_count == 3
    assert preview.series.samples[1].value is None
    assert preview.series.samples[1].quality_flag == "MISSING"
    assert len(preview.source_sha256) == 64
    assert preview.provenance["column_mapping"]["value"] == "stage"


def test_p02_model_qa_and_worker_gate_fail_closed() -> None:
    """P02: all formal runs cross the same reproducible backend and Worker gate."""

    request = _valid_qa()
    result = HydraulicModelQA().validate(request)
    assert result.run_allowed is True
    assert result.error_count == 0
    model = _frozen_model()
    digest = snapshot_hash(model.model_dump(mode="json"))
    gate = build_production_gate(request, result, model, digest)
    assert_production_gate(
        {"production_mode": True, "production_gate": gate}, model, digest
    )

    tampered = {**gate, "input_snapshot_hash": "0" * 64}
    with pytest.raises(Hydraulic1DValidationError, match="digest is invalid"):
        assert_production_gate(
            {"production_mode": True, "production_gate": tampered}, model, digest
        )

    invalid = request.model_copy(
        update={
            "engineering_crs": "EPSG:4490",
            "structures": [
                ProductionStructure(
                    structure_id="G-1",
                    structure_type="gate",
                    branch_id="1",
                    chainage_m=50,
                    vertical_datum="1985-national-height",
                    capability_status="UNSUPPORTED",
                )
            ],
        }
    )
    blocked = HydraulicModelQA().validate(invalid)
    assert blocked.run_allowed is False
    assert {item.code for item in blocked.issues} >= {
        "QA_CRS_ENGINEERING_PROJECTED_REQUIRED",
        "MODEL_ENGINE_INCOMPATIBLE",
    }


def test_p03_calibration_metrics_sweep_and_ranking_are_explicit() -> None:
    """P03: H/Q metrics, bounded sweeps, objective weights, and ranking are reproducible."""

    observed = _series("OBS-H", "water_level", [3.0, None, 3.4, 3.2])
    simulated = _series("SIM-H", "water_level", [3.1, 3.2, 3.3, 3.2])
    metrics = align_and_score(
        observed,
        simulated,
        TimeAlignmentOptions(minimum_valid_samples=3),
    )
    assert metrics.valid_sample_count == 3
    assert metrics.observed_sample_count == 3
    assert metrics.sufficient_samples
    assert metrics.rmse == pytest.approx((0.02 / 3) ** 0.5)

    plan = build_parameter_sweep(
        ParameterSweepRequest(
            parameters=[
                CalibrationParameter(
                    group_id="main-channel",
                    target_ids=["1", "2"],
                    values=[0.025, 0.03],
                )
            ],
            max_runs=2,
        )
    )
    assert plan.total_candidates == 2
    completed = [
        item.model_copy(update={"metrics": [metrics], "status": "completed"})
        for item in plan.candidates
    ]
    ranked = rank_calibration_candidates(
        CalibrationRankingRequest(
            candidates=completed,
            objective=CalibrationObjective(
                mode="water-level-focused", weights={"water_level.rmse": 1.0}
            ),
        )
    )
    assert [item.rank for item in ranked] == [1, 2]
    criteria_passed, checks = evaluate_project_metric_criteria(
        ranked[0].metrics,
        AcceptanceCriteria(
            maximum_water_level_rmse=0.1,
            minimum_observation_coverage=0.75,
        ),
    )
    assert criteria_passed
    assert all(item["passed"] for item in checks)
    criteria_failed, _ = evaluate_project_metric_criteria(
        ranked[0].metrics,
        AcceptanceCriteria(maximum_water_level_rmse=0.01),
    )
    assert not criteria_failed
    with pytest.raises(ValueError, match="must contain integers"):
        SimulationTaskCreate(
            case_id=1,
            roughness_overrides=[
                {"group_id": "invalid", "cross_section_ids": [True], "manning_n": 0.03}
            ],
        )
    with pytest.raises(ValueError, match="exceeding max_runs"):
        build_parameter_sweep(
            ParameterSweepRequest(
                parameters=[
                    CalibrationParameter(
                        group_id="g1", target_ids=["1"], values=[0.02, 0.03]
                    ),
                    CalibrationParameter(
                        group_id="g2", target_ids=["2"], values=[0.02, 0.03]
                    ),
                ],
                max_runs=3,
            )
        )


def test_p04_validation_requires_independent_evidence_and_project_criteria() -> None:
    """P04: reused calibration data cannot produce the VALIDATED model state."""

    start = datetime(2026, 7, 1, tzinfo=UTC)
    calibration = DatasetWindow(
        dataset_id="OBS-SET-01",
        event_id="FLOOD-A",
        station_ids=["STA-01"],
        start_time=start,
        end_time=start + timedelta(hours=12),
        role="calibration",
        holdout_type="independent_event",
    )
    reused = calibration.model_copy(
        update={"role": "validation", "holdout_type": "same_data"}
    )
    blocked = evaluate_validation_independence(calibration, reused)
    assert not blocked.independent
    assert blocked.issues[0].code == "VALIDATION_DATA_REUSED"

    validation = DatasetWindow(
        dataset_id="OBS-SET-02",
        event_id="FLOOD-B",
        station_ids=["STA-01"],
        start_time=start + timedelta(days=10),
        end_time=start + timedelta(days=11),
        role="validation",
        holdout_type="independent_event",
    )
    independence = evaluate_validation_independence(calibration, validation)
    metrics = align_and_score(
        _series("OBS", "water_level", [3, 3.2, 3.1]),
        _series("SIM", "water_level", [3, 3.2, 3.1]),
        TimeAlignmentOptions(),
    )
    accepted = evaluate_acceptance(
        AcceptanceEvaluationRequest(
            metrics=[metrics],
            criteria=AcceptanceCriteria(
                maximum_water_level_rmse=0.05,
                minimum_nse=0.9,
                minimum_r_squared=0.9,
            ),
            independence=independence,
        )
    )
    assert accepted.criteria_passed
    assert accepted.model_state == "VALIDATED"
    assert accepted.professional_approval_required


def test_p05_external_result_mapping_and_comparison_are_solver_neutral() -> None:
    """P05: legal external exports use explicit Branch/chainage mappings and H/Q metrics."""

    content = (
        b"reach,station,time,H,Q\n"
        b"R-A,0,0,3.0,10\nR-A,0,60,3.2,12\nR-A,0,120,3.1,11\n"
    )
    options = ExternalResultImportOptions(
        external_model_name="MIKE11",
        external_model_version="2024",
        scenario="baseline",
        vertical_datum="1985-national-height",
        time_basis="relative",
        column_mapping=ExternalColumnMapping(
            branch="reach",
            chainage="station",
            time="time",
            water_level="H",
            discharge="Q",
        ),
        branch_mappings=[
            ExternalBranchMapping(external_branch="R-A", dayu_branch="1")
        ],
    )
    preview = EngineeringDataImporter().preview_external("mike11.csv", content, options)
    assert preview.row_count == 3
    assert preview.variables == ["water_level", "discharge"]
    external = _series("EXT", "water_level", [3, 3.2, 3.1])
    dayu = _series("DAYU", "water_level", [3.05, 3.15, 3.1])
    comparison = compare_external_result(
        ExternalComparisonRequest(
            dayu_series=[dayu],
            external_series=[external],
            alignment=TimeAlignmentOptions(),
        )
    )
    assert comparison.reference_not_ground_truth
    assert comparison.metrics[0].rmse == pytest.approx((0.005 / 3) ** 0.5)
    assert len(comparison.time_series) == 3


def test_p06_result_products_exports_afflux_and_formula_safety() -> None:
    """P06: maxima, profiles, scenario deltas, afflux, and safe exports share one bundle."""

    points: list[HydraulicResultPoint] = []
    for scenario, offset in (("base", 0.0), ("project", 0.2)):
        for chainage, section in ((0.0, "=XS-1"), (100.0, "XS-2")):
            for time_seconds, water in ((0.0, 3.0), (60.0, 3.5)):
                points.append(
                    HydraulicResultPoint(
                        scenario_id=scenario,
                        branch_id="1",
                        cross_section_id=section,
                        chainage_m=chainage,
                        time_seconds=time_seconds,
                        water_level_m=water + offset,
                        discharge_m3s=10 + time_seconds / 60,
                        velocity_m_s=1.0 + time_seconds / 600,
                        depth_m=2.0 + offset,
                        bed_elevation_m=1.0,
                        geometry={"type": "Point", "coordinates": [chainage, 0]},
                    )
                )
    bundle = build_result_products(
        ResultProductRequest(
            project_id="P-SYNTHETIC",
            model_version="contract-fixture",
            baseline_scenario_id="base",
            project_scenario_id="project",
            afflux_threshold_m=0.1,
            points=points,
            calibration_table=[{"candidate_id": "reviewed-candidate", "rmse": 0.1}],
            validation_table=[{"dataset_id": "independent-event", "passed": True}],
            external_comparison_table=[{"reference": "legal-export", "difference": 0.02}],
        )
    )
    assert bundle.maximum_afflux is not None
    assert bundle.maximum_afflux["maximum_afflux_m"] == pytest.approx(0.2)
    assert bundle.afflux_reaches[0]["section_count"] == 2
    csv_bytes = export_product_csv(bundle)
    assert b"'=XS-1" in csv_bytes
    geojson = export_product_geojson(bundle)
    assert b'"FeatureCollection"' in geojson
    workbook = load_workbook(BytesIO(export_product_xlsx(bundle)), read_only=True)
    assert {
        "Summary", "Max Results", "Longitudinal Profile", "Scenario Compare",
        "Calibration", "Validation", "External Compare",
    } <= set(workbook.sheetnames)
    workbook.close()


def test_production_large_model_contract_10_branches_1000_sections_100k_points() -> None:
    """Exercise the required large synthetic result shape without claiming real-data validation."""

    points = [
        HydraulicResultPoint(
            scenario_id="large-synthetic",
            branch_id=str(branch),
            cross_section_id=f"{branch}-{section}",
            chainage_m=float(section * 100),
            time_seconds=float(step * 60),
            water_level_m=10 + branch * 0.01 + step * 0.001,
            discharge_m3s=100 + step,
            velocity_m_s=1.0 + section * 0.0001,
            depth_m=2.0,
        )
        for branch in range(10)
        for section in range(100)
        for step in range(100)
    ]
    assert len(points) == 100_000
    bundle = build_result_products(
        ResultProductRequest(
            project_id="P-LARGE-SYNTHETIC",
            model_version="performance-contract",
            project_scenario_id="large-synthetic",
            points=points,
        )
    )
    assert len(bundle.max_envelope) == 1000
    assert len(bundle.longitudinal_profile) == 1000
